from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

im = Image.open('test_block.png')

px = im.convert("RGBA").load()
print(px[3, 16])

WIDTH = im.size[0]
HEIGHT = im.size[1]
print("h", HEIGHT, "w", WIDTH)


wb = Workbook()
ws = wb.active

cells = []

def rgb2hex(r, g, b, a):
    return '{:02x}{:02x}{:02x}{:02x}'.format(r, g, b, a)



for y in range(HEIGHT):
    for x in range(WIDTH):

        a, r, g, b,  = px[x, y]
        print(rgb2hex(a, r, g, b))
        cells.append([y, x, rgb2hex(r, g, b, a).strip()])

for i in cells:
    ws.cell(i[0] + 1, i[1] + 1).fill = PatternFill('solid', fgColor=i[2])

for i in range(1, HEIGHT + 1):
    ws.row_dimensions[i].height = 3 
    

for i in range(1, WIDTH + 1):
    ws.column_dimensions[get_column_letter(i)].width = 3 
    

wb.save("test.xlsx")
